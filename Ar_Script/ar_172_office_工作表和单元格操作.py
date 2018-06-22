import openpyxl

wb=openpyxl.Workbook()

'创建三个工作表'
ws1=wb.create_sheet(title='Archer')
ws2=wb.create_sheet(title='Saber')
ws3=wb.create_sheet(title='Lancer')

'设置工作表颜色'
ws1.sheet_properties.tabColor='FF0000'
ws2.sheet_properties.tabColor='00FF00'
ws3.sheet_properties.tabColor='0000FF'

'设置第二行和第二列的行高和列宽'
ws1.row_dimensions[2].height=25
ws1.column_dimensions['B'].width=50


'合并A1到B2单元格'
ws2.merge_cells(range_string='A1:B2')
ws2['A1']='这是填充的测试内容'



wb.save(r'result/测试excel.xlsx')

# '冻结窗口'
# wb=openpyxl.load_workbook(r'resources/人口普查.xlsx')
# ws=wb.active
# ws.freenze_pens='B8'
# wb.save(r'resources/人口普查.xlsx')