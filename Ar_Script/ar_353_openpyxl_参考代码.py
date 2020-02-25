"""
openpyxl_chart_type_demos.py

（openpyxl四：图表类型）

使用：

图表类型：
#AreaChart面积图表、BarChartm条形图/柱形图/成交量图、BubbleChart气泡图、
#LineChart折线图/LineChart3D折线图、
#ScatterChart散点图、PieChart饼图/PieChart3D饼图/ProjectedPieChart投影饼图、
#StockChart股票图；

"""

####################  图表类型openpyxl.chart

######### 图表类型：
# AreaChart面积图表、BarChartm条形图/柱形图/成交量图、BubbleChart气泡图、
# LineChart折线图/LineChart3D折线图、
# ScatterChart散点图、PieChart饼图/PieChart3D饼图/ProjectedPieChart投影饼图、
# StockChart股票图

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (AreaChart,  # AreaChart面积图表（2D区域图）
                            AreaChart3D,  # AreaChart面积图表（3D区域图）

                            BarChart,  # BarChart条形图/柱形图/成交量图

                            BubbleChart,  # BubbleChart气泡图
                            LineChart,  # LineChart折线图
                            LineChart3D,  # LineChart3D折线图
                            ScatterChart,  # ScatterChart散点图
                            PieChart,  # PieChart饼图
                            PieChart3D,  # PieChart3D饼图3D
                            ProjectedPieChart,  # ProjectedPieChart投影饼图

                            StockChart,  # StockChart股票图

                            Reference,  # 用来对单元范围进行标准化引用
                            Series)  # 用来方便创建图表数据系列
from copy import deepcopy  # 深层复制
from openpyxl.chart.axis import DateAxis  # 时间轴
from openpyxl.chart.series import DataPoint  #
from openpyxl.chart.layout import (Layout,  # 布局
                                   ManualLayout  # 手动布局
                                   )

wb = Workbook()
wb_filename = r'openpyxl_chart_type_demos_case1.xlsx'

# AreaChart面积图表（2D区域图）
#####
ws_AreaChart2D = wb.create_sheet('AreaChart面积图表2D')
rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10]]
for row in rows:
    ws_AreaChart2D.append(row)
#####
# 使用：openpyxl.chart.AreaChart() 来实例化创建 2D面积图表

# 使用：Chart.title  属性 来赋值添加 图表标题
# 使用：Chart.style  属性 来赋值设置 图表样式
# 使用：Chart.x_axis.title 属性 来赋值添加 X轴标题
# 使用：Chart.y_axis.title 属性 来赋值添加 y轴标题
chart = AreaChart()  # 实例化创建面积图表对象。
chart.title = 'Area Chart'  # 图表标题
chart.style = 30  # 图表样式
chart.x_axis.title = 'Test'  # 图表X轴标题
chart.y_axis.title = 'Percentage'  # 图表Y轴标题

# 使用：openpyxl.chart.Reference(worksheet=None, min_col=None, min_row=None, max_col=None, max_row=None, range_string=None)
# 来对单元范围进行标准化引用。
labels = Reference(ws_AreaChart2D, min_col=1, min_row=1, max_row=2)  # 引用工作表上的单元范围，准备用作x轴标签使用
data = Reference(ws_AreaChart2D, min_col=2, min_row=1, max_col=3, max_row=7)  # 引用工作表上的单元范围，准备用作图表添加数据系列。

# 使用：Chart.set_categories(self, labels) 来给图表设置类别 / x轴标签。
# 使用：Chart.add_data(self, data, from_rows=False, titles_from_data=False)
# 来给图表添加一个范围的数据。默认情况下，将每个列视为一个数据系列。
chart.set_categories(labels)  # 图表设置类别/X轴标签。
chart.add_data(data, titles_from_data=True)  # 图表添加一个范围的数据。默认情况下，将每个列视为一个数据系列。

# 使用：Worksheet.add_chart(chart, anchor=None) 来在工作表单添加图表对象，至指定左上角单元格。
ws_AreaChart2D.add_chart(chart, 'A10')  # 在工作表中添加图表,可选地为左上角的锚提供单元格

# AreaChart3D面积图表（3D区域图）
#####
ws_AreaChart3D = wb.create_sheet('AreaChart面积图表3D')
for row in rows:
    ws_AreaChart3D.append(row)

#####
##使用：openpyxl.chart.AreaChart3D() 来实例化创建 3D面积图表

chart = AreaChart3D()  # 实例化创建3D面积图表对象
chart.title = 'Area Chart'  # 图表标题
chart.style = 13  # 图表样式
chart.x_axis.title = 'Test'  # 图表X轴标题
chart.y_axis.title = 'Percentage'  # 图表Y轴标题
chart.legend = None  # 图表图例

labels = Reference(ws_AreaChart3D, min_col=1, min_row=1, max_row=7)  # 引用工作表上的单元范围，准备用作x轴标签使用
data = Reference(ws_AreaChart3D, min_col=2, min_row=1, max_col=3, max_row=7)  # 引用工作表上的单元范围，准备用作图表添加数据系列。
chart.set_categories(labels)  # 图表设置类型/X轴标签
chart.add_data(data, titles_from_data=True)  # 图表添加一个范围的数据。默认情况下，将每个列视为一个数据系列。

ws_AreaChart3D.add_chart(chart, 'A10')  # 工作表单添加图表，指定左上角单元格。

# BarChart条形图和柱形图
#####
ws_BarChart = wb.create_sheet('BarChart条形图和柱形图', index=0)
rows = [
    ('Number', 'Batch1', 'Batch2'),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30)]
for row in rows:
    ws_BarChart.append(row)

#####
# 使用：openpyxl.chart.BarChart() 来实例化创建 条形图、柱形图、成交量图
# 注意：在条形图中，值被绘制为水平条或垂直列。
# 注意：通过设置type属性为col或bar，来切换垂直和水平条形图。
# 注意：如果横条是水平的，则将X和Y轴后退。
chart1 = BarChart()  # 实例化创建条形图对象
chart1.title = 'Bar Chart'  # 图表标题
chart1.style = 10  # 图表样式
chart1.type = 'col'  # 图表类型设置为’col'来定义垂直条形图
chart1.x_axis.title = 'Test Number'  # 图表X轴标题
chart1.y_axis.title = 'Sample length(mm)'  # 图表Y轴标题

labels = Reference(ws_BarChart, min_col=1, min_row=2, max_row=7)  # 引用工作表上的单元范围，作用设置X轴标签。
data = Reference(ws_BarChart, min_col=2, min_row=1, max_row=7, max_col=3)  # 引用工作表上单元范围，用作设置图表添加数据系列。
chart1.set_categories(labels)  # 图表设置类别 / X轴标签
chart1.add_data(data, titles_from_data=True)  # 图表添加一个范围的数据，默认情况下，将每个列视为一个数据系列。
chart1.shape = 4
ws_BarChart.add_chart(chart1, 'A10')  # 工作表单中添加图表，锚点左上角单元格。

# 设置水平条形图
chart2 = deepcopy(chart1)  # 深层复制 图表对象
chart2.title = 'Horizontal Bar Chart'  # 图表标题
chart2.style = 11  # 图表样式
chart2.type = 'bar'  # 注意：图表类型设置为’bar'来定义水平条形图
ws_BarChart.add_chart(chart2, 'I10')  # 工作表单中添加图表，锚点左上角单元格。

# 设置垂直、重叠条形图
# 注意：使用堆叠图表时，需要将overlap重叠属性 设置为100

# 注意：Chart.grouping 属性 分组值必须是：{'standard', 'stacked', 'clustered', 'percentStacked'}中一个。
# 用来设置重叠、百分比重叠图等共四个可选。
chart3 = deepcopy(chart1)  # 深层复制 图表对象
chart3.title = 'Stacked Chart'  # 图表标题
chart3.style = 12  # 图表样式
chart3.type = 'col'  # 注意：图表类型设置为‘col'来定义垂直条形图
chart3.grouping = 'stacked'  # 注意：图表分组属性设置为 二叠分堆放
chart3.overlap = 100  # 注意：将重叠overlap属性 设置为100 来定义使用堆叠图表。
ws_BarChart.add_chart(chart3, 'A27')

# 设置水平、重叠条形图
chart4 = deepcopy(chart1)  # 深层复制 图表对象
chart4.title = 'Percent Stacked Chart'  # 图表标题
chart4.style = 13  # 图表样式
chart4.type = 'bar'  # 图表类型
chart4.grouping = 'percentStacked'  # 注意：图表分组属性设置为 百分比+二叠分堆放
chart4.overlap = 100  # 注意：将重叠overlap属性 设置为100 来定义使用堆叠图表。
ws_BarChart.add_chart(chart4, 'I27')  # 工作表单 添加图表，锚点左上角单元格。

# BubbleChart气泡图
#####
ws_BubbleChart = wb.create_sheet('BubbleChart气泡图')
rows = [
    ('Number of Products', 'Sales in USD', 'Market share'),
    (14, 12200, 15),
    (20, 60000, 33),
    (18, 24400, 10),
    (22, 32000, 42),
    (),
    (12, 8200, 18),
    (15, 50000, 30),
    (19, 22400, 15),
    (25, 25000, 50)]
for row in rows:
    ws_BubbleChart.append(row)
#####
# 使用：openpyxl.chart.BubbleChart() 来实例化创建 BubbleChart气泡图

# 使用：Series(values, xvalues=None, zvalues=None, title=None, title_from_data=False)
# 来方便创建图表数据系列
# 参数values       : 指定 y轴值
# 参数xvalues=None : 指定 x轴值
# 参数zvalues=None : 指定 大小
# 参数title=None   : 指定图表的系列标题

# 使用：Chart.series.append(Series) 方法来添加 经过Series类 方便创建的气泡图表数据系列。

chart = BubbleChart()  # 实例化创建 气泡图表对象
chart.style = 18  # 图表样式

# 散点图表 添加第一组散点数据系列
xvalues = Reference(ws_BubbleChart, min_col=1, min_row=2, max_row=5)  # 引用数据系列，用来作为气泡的X轴坐标 即：xvalues
yvalues = Reference(ws_BubbleChart, min_col=2, min_row=2, max_row=5)  # 引用数据系列，用来作为气泡的Y轴坐标 即：values
size = Reference(ws_BubbleChart, min_col=3, min_row=2, max_row=5)  # 引用数据系列，用来作为气泡的 大小 即：zvalues
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title='2013')  # 变量赋值Series类方便创建的气泡图表数据系列。
chart.series.append(series)  # 图表系列属性 添加 经过Series类方便创建的气泡图表数据系列。

# 散点图标 添加第二组散点数据系列
xvalues = Reference(ws_BubbleChart, min_col=1, min_row=7, max_row=10)  # 引用数据系列，用来作为气泡的X轴坐标 即：xvalues
yvalues = Reference(ws_BubbleChart, min_col=2, min_row=7, max_row=10)  # 引用数据系列，用来作为气泡的Y轴坐标 即：values
size = Reference(ws_BubbleChart, min_col=3, min_row=7, max_row=10)  # 引用数据系列，用来作为气泡的 大小 即：zvalues
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title='2014')  # 变量赋值Series类方便创建的气泡图表数据系列。
chart.series.append(series)  # 图表系列属性 添加 经过Series类方便创建的气泡图表数据系列

ws_BubbleChart.add_chart(chart, 'E1')  # 工作表单 添加图表，锚点左上角单元格。

# LineChart折线图
#####
from datetime import date

ws_LineChart = wb.create_sheet('LineChart折线图')
rows = [
    ['Data', 'Batch1', 'Batch2', 'Batch3'],
    [date(2015, 9, 1), 40, 30, 25],
    [date(2015, 9, 2), 40, 25, 30],
    [date(2015, 9, 3), 50, 30, 45],
    [date(2015, 9, 4), 30, 25, 40],
    [date(2015, 9, 5), 25, 35, 30],
    [date(2015, 9, 6), 20, 40, 35]]
for row in rows:
    ws_LineChart.append(row)

#####
# 使用：openpyxl.chart.LineChart() 来实例化创建 LineChart折线图
c1 = LineChart()  # 实例化创建 折线图表实例对象
c1.title = 'Line Chart'  # 图表标题
c1.style = 13  # 图表样式
c1.x_axis.title = 'Test Number'  # 图表X轴标题
c1.y_axis.title = 'Size'  # 图表Y轴标题
data = Reference(ws_LineChart, min_col=2, max_col=4, min_row=1, max_row=7)  # 引用数据系列，用作折线图表的数据值。
c1.add_data(data, titles_from_data=True)  # 图表添加数据系列。

# 更改线型为三角线型
# 使用：Chart.series 属性 来获取所有数据系列值。
s1 = c1.series[0]  # 获取折线图表的 数据系列
s1.marker.symbol = 'triangle'  # 更改线型为三角线型
s1.marker.graphicalProperties.solidFill = 'FF0000'  # 指定固体填充
s1.marker.graphicalProperties.line.solidFill = 'FF0000'  # 指定线的固体填充
s1.graphicalProperties.line.noFill = True  # 指定是否连线。

s2 = c1.series[1]
s2.graphicalProperties.line.solifFill = '00AAAA'
s2.graphicalProperties.line.dashStyle = 'sysDot'
s2.graphicalProperties.line.width = 100050

s3 = c1.series[2]
s3.smooth = True

ws_LineChart.add_chart(c1, 'A10')

stacked = deepcopy(c1)
stacked.grouping = 'stacked'
stacked.title = 'Stacked Line Chart'  # 值必须是{' percent'， 'standard'， ' '}中的一个
ws_LineChart.add_chart(stacked, 'A27')

percent_stacked = deepcopy(c1)
percent_stacked.grouping = 'percentStacked'  # 值必须是{' percent'， 'standard'， ' '}中的一个
percent_stacked.title = 'Percent Stacked Line Chart'
ws_LineChart.add_chart(percent_stacked, 'A44')

c4 = LineChart()
c4.title = 'Date Axis'
c4.style = 12
c4.y_axis.title = 'Size'
c4.y_axis.crossAx = 500
c4.x_axis.title = 'Date'
c4.x_axis = DateAxis(crossAx=100)
c4.x_axis.number_format = 'd-mmm'
c4.x_axis.majorTimeUnit = 'days'

c4.add_data(data, titles_from_data=True)
dates = Reference(ws_LineChart, min_col=1, min_row=2, max_row=7)
c4.set_categories(dates)
ws_LineChart.add_chart(c4, 'A61')

# LineChart3D折线图3D
#####
ws_LineChart3D = wb.create_sheet('LineChart3D')
for row in rows:
    ws_LineChart3D.append(row)

#####
# 使用：openpyxl.chart.LineChart3D() 来实例化创建 LineChart3D折现图3D
c1 = LineChart3D()
c1.title = '3D Line Chart'
c1.legend = None
c1.style = 15
c1.y_axis.title = 'Size'
c1.x_axis.title = 'Test Number'

data = Reference(ws_LineChart3D, min_col=2, max_col=4, min_row=1, max_row=7)
c1.add_data(data, titles_from_data=True)
ws_LineChart3D.add_chart(c1, 'A10')

# ScatterChart散点图
#####

ws_ScatterChart = wb.create_sheet('ScatterChart散点图')
rows = [
    ['Size', 'Batch 1', 'Batch 3'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40]]
for row in rows:
    ws_ScatterChart.append(row)
#####
# 使用：openpyxl.chart.ScatterChart() 来实例化创建 ScatterChart散点图
chart = ScatterChart()
chart.title = 'Scatter Chart'
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'

xvalues = Reference(ws_ScatterChart, min_col=1, min_row=2, max_row=7)
for i in range(2, 4):
    values = Reference(ws_ScatterChart, min_col=i, min_row=1, max_row=7)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
ws_ScatterChart.add_chart(chart, 'A10')

# PieChart饼图
#####

ws_PieChart = wb.create_sheet('PieChart饼图')
data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40]]
for row in data:
    ws_PieChart.append(row)

#####
# 使用：openpyxl.chart.PieChart() 来实例化创建 PieChart饼图
pie = PieChart()
labels = Reference(ws_PieChart, min_col=1, min_row=2, max_row=5)
data = Reference(ws_PieChart, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = 'Pies sold by category'

slice = DataPoint(idx=0, explosion=20)
pie.series[0].data_points = [slice]

ws_PieChart.add_chart(pie, 'D1')

# PieChart3D饼图3D
#####
# 使用：openpyxl.chart.PieChart3D() 来实例化创建 PieChart3D饼图3D图
pie = PieChart3D()
labels = Reference(ws_PieChart, min_col=1, min_row=2, max_row=5)
data = Reference(ws_PieChart, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"

ws_PieChart.add_chart(pie, "D20")

# ProjectedPieChart投影饼图
#####
# 使用：openpyxl.chart.ProjectedPieChart() 来实例化创建 ProjectedPieChart投影饼图
ws_ProjectedPieChart = wb.create_sheet('ProjectedPieChart投影饼图')
data = [
    ['Page', 'Views'],
    ['Search', 95],
    ['Products', 4],
    ['Offers', 0.5],
    ['Sales', 0.5]]
for row in data:
    ws_ProjectedPieChart.append(row)

projected_pie = ProjectedPieChart()
projected_pie.type = 'pie'
projected_pie.splitType = 'val'
labels = Reference(ws_ProjectedPieChart, min_col=1, min_row=2, max_row=5)
data = Reference(ws_ProjectedPieChart, min_col=2, min_row=1, max_row=5)
projected_pie.add_data(data, titles_from_data=True)
projected_pie.set_categories(labels)
ws_ProjectedPieChart.add_chart(projected_pie, 'A10')

projected_bar = deepcopy(projected_pie)
projected_bar.type = 'bar'
projected_bar.splitType = 'pos'
ws_ProjectedPieChart.add_chart(projected_bar, 'A27')

# StockChart股票图
#####
from openpyxl import Workbook
from openpyxl.chart import (StockChart,  # StockChart股票图表
                            BarChart,  # BarChart条形图/柱形图/成交量图
                            Reference,  # Reference类：来对单元范围进行标准化引用。
                            Series
                            )
from openpyxl.chart.axis import (DateAxis,
                                 ChartLines  # 注意：美国线还将设置hiLoLines。K线图还需要设置upDownLines。
                                 )
from openpyxl.chart.updown_bars import UpDownBars  # 注意：美国线还将设置hiLoLines。K线图还需要设置upDownLines。

# wb=Workbook()
ws_StockChart = wb.create_sheet('StcokChart股票图', index=0)
rows = [
    ['Date', 'Valume', 'Open', 'High', 'Low', 'Close'],
    ['2015-01-01', 20000, 26.2, 27.2, 23.49, 25.45],
    ['2015-01-02', 10000, 25.45, 25.03, 19.55, 23.05],
    ['2015-01-03', 15000, 25.45, 25.03, 19.55, 23.05],
    ['2015-01-04', 2000, 22.42, 23.97, 20.07, 21.90],
    ['2015-01-05', 12000, 21.9, 23.65, 19.50, 21.51]]
for row in rows:
    ws_StockChart.append(row)

# 创建美国线股票图（高-低-收）
#####
# 使用：openpyxl.chart.StockChart() 来实例化创建 StockChart股票图表实例

# 使用：openpyxl.chart.Reference(worksheet=None, min_col=None, min_row=None, max_col=None, max_row=None, range_string=None)
# 来对单元范围进行标准化引用。

# 使用：Chart.add_data(self, data, from_rows=False, titles_from_data=False)
# 来在一次传递中添加一个范围的数据。默认情况下，将每个列视为一个数据系列。

# 使用：Chart.set_categories(self, labels) 来设置类别 / x轴值。

# 注意：高-低-收盘价 美国线基本上是没有线的折线图。
# 注意：美国线的标记设置为XYZ，还将设置hiLoLines。K线图还需要设置upDownLines。
c1 = StockChart()  # 实例化 创建 股票图表类对象
c1.title = 'High-low-close'
c1.hiLowLines = ChartLines()  # 注意：美国线的标记设置为XYZ，还将设置hiLoLines。K线图还需要设置upDownLines。

labels = Reference(ws_StockChart, min_col=1, min_row=2, max_row=6)
data = Reference(ws_StockChart, min_col=4, max_col=6, min_row=1, max_row=6)
c1.add_data(data, titles_from_data=True)
c1.set_categories(labels)

for s in c1.series:
    s.graphicalProperties.line.noFill = True
s.marker.symbol = 'dot'
s.marker.size = 5

# 注意：由于Excel中的BUG错误，仅当数据序列中的至少一个具有一些虚拟值时，才会显示高/低行。
# 可以使用以下技巧来做到这一点：
from openpyxl.chart.data_source import (NumData,
                                        NumVal)

pts = [NumVal(idx=i) for i in range(len(data) - 1)]
cache = NumData(pt=pts)
c1.series[-1].val.numRef.numCache = cache

ws_StockChart.add_chart(c1, 'A10')

# 创建K线图股票图（开-高-低-收）
#####
# 注意：开-高-低-收盘价为高-低-收盘价图表。
# 注意：美国线的标记设置为XYZ，还将设置hiLoLines。K线图还需要设置upDownLines。
c2 = StockChart()
c2.title = 'Open-hight-low-close'
c2.hiLowLines = ChartLines()  # 注意：美国线的标记设置为XYZ，还将设置hiLoLines。K线图还需要设置upDownLines。
c2.upDownBars = UpDownBars()  # 注意：美国线的标记设置为XYZ，还将设置hiLoLines。K线图还需要设置upDownLines。

labels = Reference(ws_StockChart, min_col=1, min_row=2, max_row=6)
data = Reference(ws_StockChart, min_col=3, max_col=6, min_row=1, max_row=6)
c2.add_data(data, titles_from_data=True)
c2.set_categories(labels)

for s in c2.series:
    s.graphicalProperties.line.noFill = True

# 注意：由于Excel中的BUG错误，仅当数据序列中的至少一个具有一些虚拟值时，才会显示高/低行。
# 可以使用以下技巧来做到这一点：
from openpyxl.chart.data_source import (NumData,
                                        NumVal)

pts = [NumVal(idx=i) for i in range(len(data) - 1)]
cache = NumData(pt=pts)
c2.series[-1].val.numRef.numCache = cache

ws_StockChart.add_chart(c2, 'I10')

# 创建 美国线 + 成交量 （高低收 + 成交量）
bar = BarChart()
data = Reference(ws_StockChart, min_col=2, min_row=1, max_row=6)
bar.add_data(data, titles_from_data=True)
bar.set_categories(labels)

b1 = deepcopy(bar)
c3 = deepcopy(c1)

c3.y_axis.majorGridlines = None
c3.y_axis.title = 'Price'

b1.y_axis.axId = 20
b1.z_axis = c3.y_axis
b1.y_axis.crosses = 'max'
b1 += c3

c3.title = 'High low close volume'
ws_StockChart.add_chart(b1, 'A27')

# 创建 K线图 + 成交量（开-高-低-收-量能）
b2 = deepcopy(bar)
c4 = deepcopy(c2)
c4.y_axis.majorGridlines = None
c4.y_axis.title = 'Price'
b2.y_axis.axId = 20
b2.z_axis = c4.y_axis
b2.y_axis.crosses = 'max'
b2 += c4
ws_StockChart.add_chart(b2, 'I27')

wb.save(wb_filename)






