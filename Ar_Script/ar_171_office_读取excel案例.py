import openpyxl

'读取已存在的excel文件'
wb=openpyxl.load_workbook('result/豆瓣top250.xlsx')
ws=wb.active

'用一个列表显示所有工作表的名字'
'get_sheet_names不建议使用'
names=wb.sheetnames
print(names)

# '创建工作表'
# wb.create_sheet(index=1,title='新建工作表')
# ws=wb.create_sheet(index=2,title='新建工作表2')


'复制工作表'
wb.copy_worksheet(ws)


# '删除工作表'
# wb.remove_sheet(wb.get_sheet_by_name('新建工作表'))
# wb.remove_sheet(wb.get_sheet_by_name('新建工作表2'))

# print(wb.sheetnames)

wb.save('result/豆瓣top250.xlsx')

