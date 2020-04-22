import os
import time
import openpyxl
from openpyxl.chart import LineChart,Reference,Series
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill


# app启动
class App(object):

    def __init__(self,package_info):
        self.content = ''
        self.runtime = 0
        self.beforeStart = 0
        self.afterStart = 0
        self.info=package_info

    def startApp(self):
        cmd = 'adb shell am start -W -n {}'.format(self.info[1]+"/"+self.info[2])
        self.content = os.popen(cmd).readlines()
        return self.content

    def stopApp(self):
        cmd = 'adb shell am force-stop {}'.format(self.info[1])
        os.popen(cmd)

    def clear_app_data(self):
        cmd='adb shell pm clear {}'.format(self.info[1])
        os.popen(cmd)

    def backApp(self):
        cmd = 'adb shell input keyevent 3'
        os.popen(cmd)

    def beforeAppStart(self):
        self.beforeStart = time.perf_counter()
        print(self.beforeStart)
        return self.beforeStart

    def afterAppStart(self):
        self.afterStart = time.perf_counter()
        print(self.afterStart)
        return self.afterStart

    def getTime(self):
        for line in self.content:
            if "ThisTime" in line:
                self.runtime = line.split(':')[1].strip()
                break
        return self.runtime


class Control(object):

    def __init__(self,package_info,count,driver,mode=0):
        self.app = App(package_info)
        self.count = count
        self.allData = [('runTime', 'calTime', 'currentTime')]
        self.mode=mode
        self.driver=driver

    def testOnce(self):
        before = self.app.beforeAppStart()
        self.app.startApp()
        after = self.app.afterAppStart()
        result = after - before
        time.sleep(5)
        runtime = int(self.app.getTime())

        if self.mode:
            self.app.stopApp()
        else:
            self.app.backApp()

        currentTime = self.getCurrntTime()
        self.allData.append((runtime, result, currentTime))
        time.sleep(3)

    def run(self):
        for i in range(self.count):
            self.testOnce()

    def getCurrntTime(self):
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        return currentTime

    def saveData(self,file_attr='test'):
        del self.allData[1]  # 第一行的启动数据一般不准确，先删掉

        length = len(self.allData)
        print('总行数为：',length)
        average_runtime = 0
        average_caltime = 0
        for i in range(1, length):
            average_runtime += float(self.allData[i][0])
            average_caltime += float(self.allData[i][1])

        print((average_runtime / (length - 1), average_caltime / (length - 1), self.getCurrntTime()))
        self.allData.append((average_runtime / float(length - 1), average_caltime / float(length - 1), self.getCurrntTime()))

        length = len(self.allData)

        print('总行数2为：',length)

        if os.path.exists('../test_result/app_start_time_test_data.xlsx'):
            wb=openpyxl.load_workbook('../test_result/app_start_time_test_data.xlsx')
        else:
            wb = openpyxl.Workbook()
        # sheet_num=len(wb.sheetnames)
        ws=wb.create_sheet(file_attr,index=1)
        # ws = wb.active

        '设置样式'
        font = Font(size=13, bold=True)
        alignment = Alignment(horizontal='center', vertical='center')

        for i in range(1, length+1):
            # 设置单元格的格式为数字，去掉单元格的前后空格
            ws['A{}'.format(i)].number_format = numbers.FORMAT_GENERAL
            ws['B{}'.format(i)].number_format = numbers.FORMAT_GENERAL

            ws['A{}'.format(i)] = self.allData[i -1][0]
            if i-1==0:
                ws['A{}'.format(i)].font = font
            ws['A{}'.format(i)].alignment =alignment

            ws['B{}'.format(i)] = self.allData[i-1][1]
            if i - 1 == 0:
                ws['B{}'.format(i)].font = font
            ws['B{}'.format(i)].alignment = alignment

            ws['c{}'.format(i)] = self.allData[i-1][2]
            if i - 1 == 0:
                ws['C{}'.format(i)].font = font
            ws['C{}'.format(i)].alignment =alignment

        '设置列宽'
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30

        '新增数据折线图'
        linechart = LineChart()
        linechart.title = 'Runtime Data'
        linechart.style = 39
        max_line = ws.max_row
        data = Reference(ws, min_row=1, min_col=1, max_row=max_line-1, max_col=1)
        linechart.add_data(data, titles_from_data=True)

        ws.add_chart(linechart,'E1')

        linechart2 = LineChart()
        linechart2.title = 'Caltime Data'
        linechart2.style = 39
        data2 = Reference(ws, min_row=1, min_col=2, max_row=max_line - 1, max_col=2)
        linechart2.add_data(data2, titles_from_data=True)

        ws.add_chart(linechart2, 'E19')

        wb.save('../test_result/app_start_time_test_data.xlsx')