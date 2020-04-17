import subprocess
import os
from openpyxl.chart import LineChart,Reference,Series
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill
import openpyxl
import time

def get_meminfo_data(package):

    cmd="adb shell dumpsys meminfo {}".format(package)
    info_result=subprocess.Popen(cmd,shell=True,stdout=subprocess.PIPE).stdout.readlines()
    new_info=[ str(i) for i in info_result]
    result=''
    for i in new_info:
        if 'TOTAL' in i:
            tmp='#'.join(i.split())
            result=tmp.split('#')[2]

    return result


def saveData(result_list,file_attr='test'):
    # del result_list[1]  # 第一行的启动数据一般不准确，先删掉
    cunrrent_time=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    result_list.insert(0,'启动app加载完卡片的内存值')
    length = len(result_list)

    print('总行数为：',length)


    length = len(result_list)


    if os.path.exists('test_data.xlsx'):
        wb=openpyxl.load_workbook('test_data.xlsx')
    else:
        wb = openpyxl.Workbook()
    ws=wb.create_sheet(file_attr,index=1)

    '设置样式'
    font = Font(size=13, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    for i in range(1, length+1):
        # 设置单元格的格式为数字，去掉单元格的前后空格
        ws['A{}'.format(i)].number_format = numbers.FORMAT_GENERAL
        # ws['B{}'.format(i)].number_format = numbers.FORMAT_GENERAL

        ws['A{}'.format(i)] = result_list[i-1]
        ws['A{}'.format(i)].font = font
        ws['A{}'.format(i)].alignment =alignment

        # ws['B{}'.format(i)] = result_list[i-1][1]
        # if i - 1 == 0:
        #     ws['B{}'.format(i)].font = font
        # ws['B{}'.format(i)].alignment = alignment
        #
        # ws['c{}'.format(i)] = result_list[i-1][2]
        # if i - 1 == 0:
        #     ws['C{}'.format(i)].font = font
        # ws['C{}'.format(i)].alignment =alignment

    '设置列宽'
    ws.column_dimensions['A'].width = 30
    # ws.column_dimensions['B'].width = 15
    # ws.column_dimensions['C'].width = 30

    '新增数据折线图'
    linechart = LineChart()
    linechart.title = '启动app加载完卡片的内存值'
    linechart.style = 39
    max_line = ws.max_row
    data = Reference(ws, min_row=1, min_col=1, max_row=max_line, max_col=1)
    linechart.add_data(data, titles_from_data=True)

    ws.add_chart(linechart,'C1')

    # linechart2 = LineChart()
    # linechart2.title = 'Caltime Data'
    # linechart2.style = 39
    # data2 = Reference(ws, min_row=1, min_col=2, max_row=max_line - 1, max_col=2)
    # linechart2.add_data(data2, titles_from_data=True)
    #
    # ws.add_chart(linechart2, 'E19')

    wb.save('test_data.xlsx')
    # tmp='#'.join(info_result.split())
    # for i in tmp:
    #     print(i)
    # print(tmp)

if __name__ == '__main__':
    package='com.meetu.android'
    result_list=[245305,350000,285307,266666]
    print(get_meminfo_data(package))
    saveData(result_list)