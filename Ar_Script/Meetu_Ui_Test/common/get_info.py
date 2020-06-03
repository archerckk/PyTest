import subprocess
import os
from openpyxl.chart import LineChart,Reference,Series
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill
import openpyxl
import time
import re
import easygui as g
from Ar_Script.Meetu_Ui_Test.common.app_command import *




def test_app_cpu_home_stay_cost():
    package_info=get_activity_name()
    os.popen('adb shell am start {}/{}'.format(package_info[1],package_info[2]))
    time.sleep(5)
    os.popen('adb shell input keyevent 4 ')
    result=[('测试时间','cpu百分比')]
    result.extend(get_cpu_data(package_info[1]))
    print('测试结果',result)
    data_save=Data_Save(result,'cpu_test','test_result.xlsx',(0,1),(0,'C1'))
    data_save.save_data()


def get_activity_name():
    msg = '请选择你要检查的apk安装包'
    title = '文件选择'
    default = "*.apk"
    # add_time = "_{}".format(time.localtime()[5])

    filePath = g.fileopenbox(msg=msg, title=title, default=default)

    # fileNewName = filePath.split('.apk')[0].strip() + add_time + '.apk'
    fileNewName = filePath.split('.apk')[0].strip() + '.apk'
    os.rename(filePath, fileNewName)

    print('选择的apk路径为：', fileNewName)

    command = 'aapt dumpsys badging "%s" ' % fileNewName

    handle = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
    time.sleep(2)

    reg_packageName = re.compile(r"package: name='(.+?)'")
    reg_launchableActivity = re.compile(r"launchable-activity: name='(.+?)'")

    log = str(handle.stdout.read())
    position = log.index('targetSdkVersion:')

    print('\n', log[0:position + 21], '\n')  # 截取到targetSDK Version部分包信息

    try:
        packageName = reg_packageName.search(log).group(1)
        lanuchableActivity = reg_launchableActivity.search(log).group(1).strip()
        print('选择的apk包名为：', packageName)
        print('选择的apk登录名为：', packageName+'/'+lanuchableActivity)
        return fileNewName, packageName, lanuchableActivity
    except Exception as err:
        print(err)

def get_cpu_data(package):
    result=[]
    cmd="adb shell top -n 5 -d 3 |findstr {}".format(package[0:7])
    info_result=subprocess.Popen(cmd,shell=True,stdout=subprocess.PIPE).stdout.readlines()
    # print(info_result)
    # for i in info_result:
    #     print(i)
    new_info=[ str(i) for i in info_result]
    print(new_info)
    for i in new_info:
        tmp='#'.join(i.split())
        print(tmp)
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        result.append((current_time,float(tmp.split('#')[9])))
        print(result)
    return result


def get_meminfo_data(package):

    cmd="adb shell dumpsys meminfo {}".format(package)
    info_result=subprocess.Popen(cmd,shell=True,stdout=subprocess.PIPE).stdout.readlines()
    new_info=[ str(i) for i in info_result]
    result=''
    for i in new_info:
        if 'TOTAL' in i:
            tmp='#'.join(i.split())
            result=int(tmp.split('#')[2])

    return result


def saveData(result_list,file_attr='test'):
    # del result_list[1]  # 第一行的启动数据一般不准确，先删掉
    cunrrent_time=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    result_list.insert(0,'启动app加载完卡片的内存值')
    length = len(result_list)

    print('总行数为：',length)


    length = len(result_list)


    if os.path.exists('../test_result/meminfo_test_data.xlsx'):
        wb=openpyxl.load_workbook('../test_result/meminfo_test_data.xlsx')
    else:
        wb = openpyxl.Workbook()
    ws=wb.create_sheet(file_attr,index=1)

    '设置样式'
    font = Font(size=13, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    for i in range(1, length+1):
        # 设置单元格的格式为数字，去掉单元格的前后空格
        ws['A{}'.format(i)].number_format = numbers.FORMAT_GENERAL
        # ws['B{}'.format(i)].number_format = numbers.FORMAT_GENERAL

        ws['A{}'.format(i)] = result_list[i-1]
        ws['A{}'.format(i)].font = font
        ws['A{}'.format(i)].alignment =alignment

        # ws['B{}'.format(i)] = result_list[i-1][1]
        # if i - 1 == 0:
        #     ws['B{}'.format(i)].font = font
        # ws['B{}'.format(i)].alignment = alignment
        #
        # ws['c{}'.format(i)] = result_list[i-1][2]
        # if i - 1 == 0:
        #     ws['C{}'.format(i)].font = font
        # ws['C{}'.format(i)].alignment =alignment

    '设置列宽'
    ws.column_dimensions['A'].width = 30
    # ws.column_dimensions['B'].width = 15
    # ws.column_dimensions['C'].width = 30

    '新增数据折线图'
    linechart = LineChart()
    linechart.title = '启动app加载完卡片的内存值'
    linechart.style = 39
    max_line = ws.max_row
    data = Reference(ws, min_row=1, min_col=1, max_row=max_line, max_col=1)
    linechart.add_data(data, titles_from_data=True)

    ws.add_chart(linechart,'C1')

    # linechart2 = LineChart()
    # linechart2.title = 'Caltime Data'
    # linechart2.style = 39
    # data2 = Reference(ws, min_row=1, min_col=2, max_row=max_line - 1, max_col=2)
    # linechart2.add_data(data2, titles_from_data=True)
    #
    # ws.add_chart(linechart2, 'E19')

    wb.save('../test_result/meminfo_test_data.xlsx')
    # tmp='#'.join(info_result.split())
    # for i in tmp:
    #     print(i)
    # print(tmp)

if __name__ == '__main__':
    package='com.real.app.android'
    result_list=[245305,350000,285307,266666]
    # print(get_cpu_data(package))
    # print(get_meminfo_data(package))
    test_app_cpu_home_stay_cost()
    # saveData(result_list)