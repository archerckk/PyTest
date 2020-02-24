import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill
import time
from openpyxl.styles.colors import RED


def url_open(url):

    header= {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3904.108 Safari/537.36',
             'Upgrade-Insecure-Requests': 1}
             # 'cookie':'ll="118281"; bid=0Mbo_K31YXs; __gads=ID=8f4d6e5428f23ef7:T={}:S=ALNI_MaoXmCmyXgw2Az86gxoFZpCsuHvMQ; __yadk_uid=F9cjRb4y2FSfmODEhZCOv0hznJe3s43b; _vwo_uuid_v2=DC8825503EF0B1F64A605B5CC19B330D9|8372ec25fdb63dc67c8d40c1a6bfc94f; ap_v=0,6.0; __utmz=30149280.1582370178.2.2.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; __utmz=223695111.1582370178.2.2.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; __utma=30149280.401442872.1580737379.1582370178.1582375861.3; __utmc=30149280; __utma=223695111.48937608.1580737379.1582370178.1582375861.3; __utmb=223695111.0.10.1582375861; __utmc=223695111; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1582375861%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DPmTyBmrzLb889XXcwGBfzmS9dOsNpgWlc4W4QSkIakURSZLrff0sw-taAG-KhaPWwNWhTJQE65GKSqfEdeE_Ca%26wd%3D%26eqid%3Debb79de10004d0fa000000055e510d7a%22%5D; _pk_ses.100001.4cf6=*; __utmt_t1=1; __utmb=30149280.9.8.1582376530926; _pk_id.100001.4cf6=a726c18ea1c59195.1580737379.3.1582376531.1582370179.; RT=s=1582376579758&r=https%3A%2F%2Fmovie.douban.com%2Ftop250'.format(int(time.time()))}
    proxy={'http':'http://221.214.181.98:53281'}
    res=requests.get(url,header,proxies=proxy)

    return res


def find_movies(res):
    soup = bs(res.text, 'html.parser')

    '电影名'
    movies=[]
    targets=soup.find_all('div',class_='hd')
    for i in targets:
        movies.append(i.a.span.text+'  ')


    '评分'
    score=[]
    targets = soup.find_all('span', class_='rating_num')
    for i in targets:
        #txt执行代码
        # score.append('评分：%s  '%i.text)
        'excel执行代码'
        score.append(i.text)


    '电影介绍'
    info=[]
    targets=soup.find_all('div',class_='bd')
    for i in targets:
        try:
            info.append(i.p.text.split('\n')[1].strip()+i.p.text.split('\n')[2].strip())
        except:
            continue


    '显示结果'
    result=[]
    length=len(movies)
    for  i in range(length):
        # 'txt执行代码'
        # result.append(movies[i]+score[i]+info[i]+'\n')
        'excel执行代码'
        result.append([movies[i],score[i],info[i]])


    return result




def find_depth(res):
    soup = bs(res.text, 'html.parser')
    depth=soup.find('span',class_='next').previous_sibling.previous_sibling.text

    return int(depth)


def save_excel(result):
    wb=openpyxl.Workbook()
    wb.guess_type=True
    ws=wb.active


    '标题样式'
    title_style=NamedStyle(name='title_style')
    title_style.alignment=Alignment(horizontal='center',vertical='center')
    title_style.font=Font(size=14,bold=True)
    title_style.fill=PatternFill(fill_type='solid',fgColor='b2b2b2')


    '内容样式'
    content_style = NamedStyle(name='content_style')
    content_style.alignment = Alignment(horizontal='center', vertical='center')
    content_style.font = Font(size=12)

    '注册标题和内容样式'
    wb.add_named_style(title_style)
    wb.add_named_style(content_style)

    ws.append(['电影名','评分','电影介绍'])
    for i in ws.rows:

        ws[i[0].coordinate].style=title_style
        ws.column_dimensions['A'].width=20
        ws[i[1].coordinate].style = title_style
        ws.column_dimensions['B'].width = 20
        ws[i[2].coordinate].style = title_style
        ws.column_dimensions['C'].width = 150


    for i in result:
        ws.append(i)


    for i in ws.iter_rows(min_row=2,min_col=1,max_row=251,max_col=3):
        ws[i[0].coordinate].style=content_style
        ws.column_dimensions['A'].width = 20
        # ws[i[1].coordinate].number_format = '[>9.5][RED]#.0分'
        ws[i[1].coordinate].style = content_style
        ws.column_dimensions['B'].width = 20
        ws[i[2].coordinate].style = content_style
        ws.column_dimensions['C'].width = 150




    wb.save('result/豆瓣top250.xlsx')



def main():
    host='https://movie.douban.com/top250'
    res=url_open(host)
    print(res.status_code)

    '获取一个要爬取的页面总数'
    depth=find_depth(res)

    result=[]
    for i in range(depth):
        url=host+'?start=%d&filter='%(25*i)
        res=url_open(url)
        result.extend(find_movies(res))
        time.sleep(5)

    # 'txt执行代码'
    # with open('result/douban.txt','w',encoding='utf-8')as f:
    #     f.writelines(result)

    'excel执行代码'
    save_excel(result)



if __name__ == '__main__':
    main()