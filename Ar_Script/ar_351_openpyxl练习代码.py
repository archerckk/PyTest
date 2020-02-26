import openpyxl
from openpyxl.chart import Series
from openpyxl.styles import Font,Alignment,numbers
from openpyxl.chart import LineChart,Reference
import random
# wb=openpyxl.load_workbook('test_data.xlsx')

wb=openpyxl.load_workbook('test_code.xlsx')
# wb=openpyxl.Workbook()
ws=wb.create_sheet('test1_{}'.format(random.randint(1,100)),0)
# ws_list=wb.sheetnames
# print(ws_list)
# ws=wb['数值说明']

title_list=['标题1','标题2','标题3']
value_list=[(5,11,6),(7,15,4),(22,11,15)]
#
font_title=Font(size=13,bold=True)
alignment=Alignment(horizontal='center',vertical='center')

#
ws.append(title_list)
#
for row in ws.iter_rows():
    for i in row:
        i.font=font_title
        i.alignment=alignment

for i in value_list:
    ws.append(i)

for row in ws.iter_rows():
    for i in row:
        i.alignment=alignment
        i.number_format=numbers.FORMAT_GENERAL
#3,10
# print("新增测试代码")
linechart=LineChart()
linechart.title='数据折线图'
linechart.style=39
max_line=ws.max_row
data=Reference(ws,min_row=1,min_col=1,max_row=max_line,max_col=3)
linechart.add_data(data,titles_from_data=True)


# series_obj = linechart.series[2]

# linechart.append(series_obj)

ws.add_chart(linechart, 'E1')

wb.save('test_code.xlsx')