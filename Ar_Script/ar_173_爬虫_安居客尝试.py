import requests
import bs4
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import re


def url_open(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'}
    res = requests.get(url=url, headers=headers)

    return res


def info_find(res):
    # '测试代码'
    # with open('result/安居客.txt','r',encoding='utf-8')as f:
    #     soup=bs4.BeautifulSoup(f.read(),'html.parser')

    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    '房产编码获取'
    code = []
    url_list = []
    targets = soup.find_all('div', class_='house-title')

    for i in targets:
        url_list.append(i.a['href'])

    for i in url_list:
        res = url_open(i)
        soup2 = bs4.BeautifulSoup(res.text, 'html.parser')
        targets = soup2.find_all('span', class_='house-encode')
        for i in targets:
            code.append(i.text)
    code_new = []
    for i in code:
        i = i.split('，')[0].strip()
        i = i.split('：')[1].strip()
        # print(i)
        code_new.append(i)

    '标题信息筛选'
    title = []
    targets = soup.find_all('div', class_='house-title')
    for i in targets:
        title.append(i.a.text.strip())

    '总价信息'
    total = []
    targets = soup.find_all('span', class_='price-det')
    for i in targets:
        total.append(i.text)

    '单价信息'
    price = []
    targets = soup.find_all('span', class_='unit-price')
    for i in targets:
        price.append(i.text)

    '详细信息'
    detail = []
    targets = soup.find_all('div', class_='details-item')
    for i in targets:
        if '' in i.text.strip():
            detail.append(i.text.strip().split('')[0])

    '地址信息'
    address = []
    targets = soup.find_all('span', class_='comm-address')
    for i in targets:
        i = i.text.strip().split('\n')
        a = str(i[0].split()).strip(r'[]').strip("'")
        b = i[1].strip()
        address.append(a + '——' + b)

    result = []
    length = len(code)

    for i in range(length):
        result.append([code_new[i], title[i], total[i], price[i], detail[i], address[i]])

    # length = len(result)
    # print(length)

    result_dict = {}

    for i in result:
        try:
            result_dict[i[0]] = i
        except KeyError:
            result_dict[i[0]] = i

    # for i in result_dict.items():
    #     print(i[0], i[1])

    result_final = []
    for i in result_dict.keys():
        result_final.append(result_dict[i])

    return result_final


def compare(result, length):
    wb = openpyxl.load_workbook('result/安居客.xlsx')
    ws = wb.active
    # result[' 1294849477'][3] = '8000元/m²'
    # result[' 1285562934'][3] = '8000元/m²'
    for i in ws.iter_rows(min_row=2, min_col=1, max_row=length, max_col=6):
        if (str(i[0].value)) in result:
            new = int(re.search(r'\d+', result[str(i[0].value).strip()][3]).group())
            old = int(re.search(r'\d+', i[3].value).group())
            differ = 0
            if new < old:
                differ = old - new
                # print('单价下降：%d' % differ)
                result[str(i[0].value).strip()][4] = ('-%d' % differ)
                # print(result[str(i[0].value).strip()])
            elif new > old:
                differ = new - old
                # print('单价上升：%d' % differ)
                result[str(i[0].value).strip()][4] = ('+%d' % differ)
            elif new == old:
                result[str(i[0].value).strip()][4] = '0'

    # for i in result:
    #     print(result[i])


    return result


def save_to_excel(result):
    wb = openpyxl.Workbook()
    ws = wb.active

    '标题样式'
    title_style = NamedStyle(name='title_style')
    title_style.font = Font(size='16', bold=True)
    title_style.alignment = Alignment(horizontal='center', vertical='center')
    title_style.fill = PatternFill(fill_type='solid', fgColor='b2b2b2')

    '内容样式'
    content_style = NamedStyle(name='content_style')
    content_style.font = Font(size='12')
    content_style.alignment = Alignment(horizontal='center', vertical='center')

    ws.add_named_style = title_style

    ws.append(['房产编码', '标题', '总价', '单价', '浮动', '详细信息', '地址'])

    for i in ws.rows:
        ws[i[0].coordinate].style = title_style
        ws.column_dimensions['A'].width = 15
        ws[i[1].coordinate].style = title_style
        ws.column_dimensions['B'].width = 80
        ws[i[2].coordinate].style = title_style
        ws.column_dimensions['C'].width = 10
        ws[i[3].coordinate].style = title_style
        ws.column_dimensions['D'].width = 10
        ws[i[4].coordinate].style = title_style
        ws.column_dimensions['E'].width = 15
        ws[i[5].coordinate].style = title_style
        ws.column_dimensions['F'].width = 60
        ws[i[6].coordinate].style = title_style
        ws.column_dimensions['G'].width = 60

    for i in result.keys():
        ws.append(result[i])

    length = len(result)
    for i in ws.iter_rows(min_row=2, min_col=1, max_row=length + 1, max_col=7):
        ws[i[0].coordinate].style = content_style
        ws.column_dimensions['A'].width = 15
        ws[i[1].coordinate].style = content_style
        ws.column_dimensions['B'].width = 80
        ws[i[2].coordinate].style = content_style
        ws.column_dimensions['C'].width = 10
        ws[i[3].coordinate].style = content_style
        ws.column_dimensions['D'].width = 10
        ws[i[4].coordinate].style = content_style
        ws.column_dimensions['E'].width = 15
        ws[i[5].coordinate].style = content_style
        ws.column_dimensions['F'].width = 60
        ws[i[6].coordinate].style = content_style
        ws.column_dimensions['G'].width = 60

    wb.save('result/安居客.xlsx')


def main():
    length = 1
    result = []

    # result=[['1294849477', '时代倾城 70万 3室2厅1卫 精装修，难得的好户型急售', '70万', '8000元/m²', '3室2厅|81m²|高层(共33层)|2013年建造',
    #    '时代倾城——清城-清城-大学西路222号'],
    #         ['1285562934', '时代倾城 71万 3室2厅1卫 精装修，难得的好户型急售', '70万', '7000元/m²', '3室2厅|81m²|高层(共33层)|2013年建造',
    #          '时代倾城——清城-清城-大学西路222号'],
    #         ['1282902691', '时代倾城 71万 3室2厅1卫 精装修，难得的好户型急售', '70万', '9243元/m²', '3室2厅|81m²|高层(共33层)|2013年建造',
    #          '时代倾城——清城-清城-大学西路222号'],
    #         ['1294849404', '时代倾城 71万 3室2厅1卫 精装修，难得的好户型急售', '70万', '9666元/m²', '3室2厅|81m²|高层(共33层)|2013年建造',
    #          '时代倾城——清城-清城-大学西路222号']
    #         ]

    for i in range(length):
        host = 'https://qingyuan.anjuke.com/sale/p{}-rd1/?kw=时代倾城#filtersort'.format(i + 1)
        res = url_open(host)
        # print(res.text)
        result.extend(info_find(res))

    # print(len(result))
    # for i in result:
    #     print(i)

    result_dict = {}
    for i in result:
        result_dict[i[0]] = i

    result_length = len(result_dict)
    # result_length = 180

    for i in result_dict:
        result_dict[i].insert(4, '0')

    # for i in result_dict:
    #     print()

    # compare(result_dict, result_length)
    result_final = compare(result_dict, result_length)

    for i in result_final:
        print(result_final[i])

    save_to_excel(result_final)


if __name__ == '__main__':
    main()
