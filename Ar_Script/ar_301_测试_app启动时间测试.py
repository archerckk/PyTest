import os
import re
import time
import csv
import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill
import subprocess
import easygui as g
import random

"""
caltime:启动app后的计时器的值-启动app前启动计时器的值
runtime:adb shell am start -W -n 获取的total时间值
currentTime:测试执行时间
最后一行为平均值
"""


'''
合并csv文件到一个excel文档

1、打开一个空的excel文件
2、打开csv文件，将里面的内容复制到excel文件里面
'''


def get_activity_name():
    msg = '请选择你要检查的apk安装包'
    title = '文件选择'
    default = "*.apk"
    add_time = "_{}".format(time.localtime()[5])

    filePath = g.fileopenbox(msg=msg, title=title, default=default)

    fileNewName = filePath.split('.apk')[0].strip() + add_time + '.apk'
    os.rename(filePath, fileNewName)

    print('选择的apk路径为：', fileNewName)

    command = 'aapt dumpsys badging "%s" ' % fileNewName

    handle = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
    time.sleep(2)

    reg_packageName = re.compile(r"package: name='(.+?)'")
    reg_launchableActivity = re.compile(r"launchable-activity: name='(.+?)'")

    log = str(handle.stdout.read())
    position = log.index('targetSdkVersion:')

    print('\n', log[0:position + 21], '\n')  # 截取到targetSDK Version部分包信息

    try:
        packageName = reg_packageName.search(log).group(1)
        lanuchableActivity = reg_launchableActivity.search(log).group(1).strip()
        print('选择的apk包名为：', packageName)
        print('选择的apk登录名为：', packageName+'/'+lanuchableActivity)
        return fileNewName, packageName, lanuchableActivity
    except Exception as err:
        print(err)


# app启动
class App(object):

    def __init__(self,package_info):
        self.content = ''
        self.runtime = 0
        self.beforeStart = 0
        self.afterStart = 0
        self.info=package_info

    def startApp(self):
        cmd = 'adb shell am start -W -n {}'.format(self.info[1]+"/"+self.info[2])
        self.content = os.popen(cmd).readlines()
        return self.content

    def stopApp(self):
        cmd = 'adb shell am force-stop {}'.format(self.info[1])
        os.popen(cmd)

    def backApp(self):
        cmd = 'adb shell input keyevent 3'
        os.popen(cmd)

    def beforeAppStart(self):
        self.beforeStart = time.perf_counter()
        print(self.beforeStart)
        return self.beforeStart

    def afterAppStart(self):
        self.afterStart = time.perf_counter()
        print(self.afterStart)
        return self.afterStart

    def getTime(self):
        for line in self.content:
            if "ThisTime" in line:
                self.runtime = line.split(':')[1].strip()
                break
        return self.runtime


class Control(object):

    def __init__(self,package_info,count,mode=0,):
        self.app = App(package_info)
        self.count = count
        self.allData = [('runTime', 'calTime', 'currentTime')]
        self.mode=mode

    def testOnce(self):
        before = self.app.beforeAppStart()
        self.app.startApp()
        after = self.app.afterAppStart()
        result = after - before
        time.sleep(5)
        runtime = self.app.getTime()

        if self.mode:
            self.app.stopApp()

        self.app.backApp()
        currentTime = self.getCurrntTime()
        self.allData.append((runtime, result, currentTime))
        time.sleep(3)

    def run(self):
        for i in range(self.count):
            self.testOnce()

    def getCurrntTime(self):
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        return currentTime

    def saveData(self,file_attr='test'):
        del self.allData[1]  # 第一行的启动数据一般不准确，先删掉

        length = len(self.allData)
        print('总行数为：',length)
        average_runtime = 0
        average_caltime = 0
        for i in range(1, length):
            average_runtime += float(self.allData[i][0])
            average_caltime += float(self.allData[i][1])

        print((average_runtime / (length - 1), average_caltime / (length - 1), self.getCurrntTime()))
        self.allData.append((average_runtime / float(length - 1), average_caltime / float(length - 1), self.getCurrntTime()))

        length = len(self.allData)

        print('总行数2为：',length)

        if os.path.exists('test_data.xlsx'):
            wb=openpyxl.load_workbook('test_data.xlsx')
        else:
            wb = openpyxl.Workbook()
        sheet_num=len(wb.sheetnames)
        ws=wb.create_sheet(file_attr,index=sheet_num+1)
        # ws = wb.active

        '设置样式'
        font = Font(size=13, bold=True)
        alignment = Alignment(horizontal='center', vertical='center')

        for i in range(1, length+1):
            # 设置单元格的格式为数字，去掉单元格的前后空格
            ws['A{}'.format(i)].number_format = numbers.FORMAT_GENERAL
            ws['B{}'.format(i)].number_format = numbers.FORMAT_GENERAL

            ws['A{}'.format(i)] = self.allData[i -1][0]
            if i-1==0:
                ws['A{}'.format(i)].font = font
            ws['A{}'.format(i)].alignment =alignment

            ws['B{}'.format(i)] = self.allData[i-1][1]
            if i - 1 == 0:
                ws['B{}'.format(i)].font = font
            ws['B{}'.format(i)].alignment = alignment

            ws['c{}'.format(i)] = self.allData[i-1][2]
            if i - 1 == 0:
                ws['C{}'.format(i)].font = font
            ws['C{}'.format(i)].alignment =alignment

        '设置列宽'
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30

        wb.save('test_data.xlsx')


        #csv 保存
        # csvFile=open('testData_{}.csv'.format(file_attr),'w',newline='')
        # writer=csv.writer(csvFile)
        #
        # writer.writerows(self.allData)
        #
        # csvFile.close()







if __name__ == '__main__':
    package_info=get_activity_name()
    # package_info=(" ",'com.tinder','com.tinder.activities.LoginActivity')
    control = Control(package_info,16,mode=1)
    control.run()
    control.saveData('冷启动_v10_{}'.format(random.randint(1,100)))

    control = Control(package_info,16)
    control.run()
    control.saveData('热启动_v10_{}'.format(random.randint(1,100)))

